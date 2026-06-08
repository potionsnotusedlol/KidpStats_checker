from enum import IntEnum
from config import config
from openpyxl import load_workbook
from datetime import date

import aiosqlite
import asyncio

ROLES_FILE_PATH = config.STORAGE_FOLDER.get_secret_value() + config.ROLES_FILENAME.get_secret_value()
ROLES_DB_PATH = config.STORAGE_FOLDER.get_secret_value() + config.ROLES_DB_NAME.get_secret_value()
INFO_FILE_PATH = config.STORAGE_FOLDER.get_secret_value() + config.INFO_FILENAME.get_secret_value()
INFO_DB_PATH = config.STORAGE_FOLDER.get_secret_value() + config.INFO_DB_NAME.get_secret_value()
MAX_HISTORICAL_STORE = 10

# region DB handling
async def get_user_db() -> None:
    """
    Function to fetch roles usernames to and from DB. Creates one if no DB found.

    :return: Returns :type:`None`, since the output not used in any other function.
    """

    loop = asyncio.get_event_loop()
    workbook = await loop.run_in_executor(None, lambda: load_workbook(filename=ROLES_FILE_PATH, read_only=True))
    roles_file = workbook.active
    rows = []

    for row in roles_file.iter_rows(min_row=2, max_col=2, values_only=True): # type: ignore
        fullname, telegram_id = row[0], row[1]

        if not telegram_id or not fullname:
            continue
        elif telegram_id.strip("@") == config.OWNER_USERNAME.get_secret_value(): # type: ignore
            continue

        rows.append((str(telegram_id).strip("@"), str(fullname).strip()))
    
    workbook.close()

    async with aiosqlite.connect(ROLES_DB_PATH) as role_database:
        await role_database.execute(
            """
                CREATE TABLE IF NOT EXISTS Users (
                    telegram_ID TEXT PRIMARY KEY,
                    fullname TEXT NOT NULL
                )
        """)
        await role_database.executemany(
            """
                INSERT INTO Users (telegram_ID, fullname)
                VALUES (?, ?)
                ON CONFLICT (telegram_ID) DO UPDATE SET fullname = excluded.fullname
            """,
            rows
        )

        await role_database.commit()

async def fetch_info_file(filename: str) -> None:
    """
    Fetches info from .xlsx file into DB.

    :param filname: Name of the file containing the information
    :return: Returns :type:`None`, just translates info into DB.
    """

    loop = asyncio.get_event_loop()
    wb = await loop.run_in_executor(None, lambda: load_workbook(filename, read_only=True))

    for sheet in wb.worksheets:
        table_name = filename[50:-5] + "_" + sheet.title # change the filename's slice later
        table_name = table_name.replace("-", "_")

        async with aiosqlite.connect(INFO_DB_PATH) as info_db:
            await info_db.execute(
                """
                    CREATE TABLE IF NOT EXISTS {} (
                        student_fullname TEXT PRIMARY KEY,
                        pptx TEXT,
                        docx TEXT,
                        text TEXT,
                        readme TEXT,
                        optional TEXT,
                        work INTEGER,
                        dev INTEGER,
                        size TEXT,
                        last_load TEXT,
                        before_loaded TEXT,
                        milestone TEXT,
                        out_milestone TEXT,
                        active TEXT,
                        closed TEXT,
                        total_issues TEXT,
                        missed TEXT,
                        progress TEXT,
                        theme TEXT,
                        objective TEXT,
                        problems TEXT,
                        literature TEXT,
                        recommendations TEXT,
                        supervisor TEXT,
                        consultant TEXT
                    )
                """.format(table_name),
            )
            await info_db.commit()
        
            rows = []

            for row in sheet.iter_rows(min_row=2):
                data = tuple()

                for item in row:
                    data += (item.value,)
            
                rows.append(data)
        
            await info_db.executemany(
                f"""
                    INSERT OR REPLACE INTO {table_name}
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                rows
            )
            await info_db.commit()

async def cleanup_data_DB() -> None:
    """
    Function to regulate maximum amount of historical records for main DB.

    :return: Returns :type:`None`, since the action is atomic.
    """

    async with aiosqlite.connect(INFO_DB_PATH) as info_db:
        async with info_db.execute(
            """
                SELECT COUNT(*)
                FROM sqlite_master
                WHERE type='table'
            """
        ) as cursor:
            count = await cursor.fetchone()

        if not count or count <= MAX_HISTORICAL_STORE * 6:
            return
        else:
            # pattern = re.compile(rf"^_\d{{4}}_\d{{2}}_\d{{2}}_$")

            async with info_db.execute(
                """
                    SELECT *
                    FROM sqlite_master
                    WHERE type='table'
                """
            ) as cursor:
                db_names = await cursor.fetchall()
                dates_str = set([str(name[1:11]) for name in db_names])
                dates = list()

                for date_str in dates_str:
                    dates.append(date(int(x[0]), int(x[2]), int(x[1])) for x in date_str.split("_"))
# endregion

class Request(IntEnum):
    """
    Requests for Info handling and module usage. Class to be more convenient.
    """

    UPDATE_ROLES_DB = 0
    UPDATE_DATA_DB = 1

async def SDH(request: Request, filename: str | None = None) -> None:
    """
    Delivers communications to this module from other program parts.

    :param request: :class:`Request` option to tell the module what to do
    :param filename: Optional :type:`str` for functions that require passing :code:`filename` as an argument
    :return: Returns :type:`None`, no output needed.
    """

    if request == Request.UPDATE_ROLES_DB:
        await get_user_db()
    elif request == Request.UPDATE_DATA_DB:
        await fetch_info_file(filename) # type: ignore