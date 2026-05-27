from enum import IntEnum
from Config import config
from openpyxl import load_workbook

import aiosqlite
import asyncio

ROLES_FILE_PATH = config.STORAGE_FOLDER.get_secret_value() + config.ROLES_FILENAME.get_secret_value()
ROLES_DB_PATH = config.STORAGE_FOLDER.get_secret_value() + config.ROLES_DB_NAME.get_secret_value()
INFO_FILE_PATH = config.STORAGE_FOLDER.get_secret_value() + config.INFO_FILENAME.get_secret_value()
INFO_DB_PATH = config.STORAGE_FOLDER.get_secret_value() + config.INFO_DB_NAME.get_secret_value()

# region DB handling
async def getUsersDB():
    loop = asyncio.get_event_loop()
    workbook = await loop.run_in_executor(None, lambda: load_workbook(filename=ROLES_FILE_PATH, read_only=True))
    roles_file = workbook.active
    rows = []

    for row in roles_file.iter_rows(min_row=2, max_col=2, values_only=True): # type: ignore
        fullname, telegram_id = row[0], row[1]

        if not telegram_id or not fullname:
            continue
        elif telegram_id.strip("@") == config.OWNER_USERNAME.get_secret_value(): # type: ignore
            continue

        rows.append((str(telegram_id).strip("@"), str(fullname).strip()))
    
    workbook.close()

    async with aiosqlite.connect(ROLES_DB_PATH) as role_database:
        await role_database.execute(
            """
                CREATE TABLE IF NOT EXISTS Users (
                    telegram_ID TEXT PRIMARY KEY,
                    fullname TEXT NOT NULL
                )
        """)
        await role_database.executemany(
            """
                INSERT INTO Users (telegram_ID, fullname)
                VALUES (?, ?)
                ON CONFLICT (telegram_ID) DO UPDATE SET fullname = excluded.fullname
            """,
            rows
        )

        await role_database.commit()

async def updateInfoFile():
    pass

async def fetchInfoFile(filename):
    loop = asyncio.get_event_loop()
    wb = await loop.run_in_executor(None, lambda: load_workbook(filename, read_only=True))

    for sheet in wb.worksheets:
        table_name = filename[50:-5] + "_" + sheet.title # change the filename's slice later
        table_name = table_name.replace("-", "_")

        async with aiosqlite.connect(INFO_DB_PATH) as info_db:
            await info_db.execute(
                """
                    CREATE TABLE IF NOT EXISTS {}(
                        student_fullname TEXT PRIMARY KEY,
                        pptx TEXT,
                        docx TEXT,
                        text TEXT,
                        readme TEXT,
                        optional TEXT,
                        work INTEGER,
                        dev INTEGER,
                        size TEXT,
                        last_load TEXT,
                        before_loaded TEXT,
                        milestone TEXT,
                        out_milestone TEXT,
                        active TEXT,
                        closed TEXT,
                        total_issues TEXT,
                        missed TEXT,
                        progress TEXT,
                        theme TEXT,
                        objective TEXT,
                        problems TEXT,
                        literature TEXT,
                        recommendations TEXT,
                        supervisor TEXT,
                        consultant TEXT
                    )
                """.format(table_name),
            )
            await info_db.commit()
        
            rows = []

            for row in sheet.iter_rows(min_row=2):
                data = tuple()

                for item in row:
                    data += (item.value,)
            
                rows.append(data)
        
            await info_db.executemany(
                f"""
                    INSERT OR REPLACE INTO {table_name}
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                rows
            )
            await info_db.commit()



# endregion

class Request(IntEnum):
    UPDATE_ROLES_DB = 0
    UPDATE_DATA_DB = 1

async def SDH(request: Request):
    if request == Request.UPDATE_ROLES_DB:
        await getUsersDB()
    elif request == Request.UPDATE_DATA_DB:
        pass